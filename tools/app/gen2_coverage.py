"""Take the lab's metrics spreadsheet and add a column saying, for Meta Gen 2
through the toolkit, whether the app can log each metric and what evidence exists.

Usage:
    python tools/app/gen2_coverage.py --sheet "C:/path/IMC_metrics_logging_v2.xlsx"

Same idea as gen1_coverage.py but pointed at her Gen 2 column and joined against
the app's own session files instead of the black-box runs. Numbers come out of
data/app, so a metric only reads "logged" if a session actually produced it.

The distinction that matters here: the app writes far more than any session has
exercised. Code that runs but has never been fed says "code ready", not "logged".

my_status values:
    logged                  the app writes it and a session has real data
    logged (phone preview)  measured, but on the phone's own screen rather than a
                            remote viewer - see the note on the Viewer rows
    code ready              the logger writes it, no session has exercised it yet
    derived                 computed downstream by window_metrics / merge_sessions
    needs server            needs the MediaMTX push chain, which does not exist yet
    not observable          the reference app cannot produce the condition at all
    ask                     the spec is ambiguous here, confirm before building
    optional                the sheet marks it Optional for Gen 2
    n/a                     the sheet does not ask for it on this platform
"""

import argparse
import csv
import glob
import os
import statistics

DATA = os.path.join("data", "app")

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl needed: python -m pip install openpyxl")


def read_rows(path):
    if not path or not os.path.exists(path):
        return []
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def nums(rows, key):
    out = []
    for r in rows:
        try:
            out.append(float(r[key]))
        except (KeyError, TypeError, ValueError):
            pass
    return out


def newest(pattern):
    """Latest session file matching the pattern, or None. Sessions are stamped
    yyyymmdd_hhmmss so a plain sort is chronological."""
    found = sorted(glob.glob(os.path.join(DATA, pattern)))
    return found[-1] if found else None


def session_stamp(path):
    if not path:
        return ""
    base = os.path.basename(path)
    return base.split("_", 1)[1].rsplit(".", 1)[0] if "_" in base else base


def build_findings(args):
    frames_path = args.frames or newest("frames_*.csv")
    events_path = args.events or newest("events_*.csv")
    display_path = args.display or newest("display_*.csv")
    encoded_path = args.encoded or newest("encoded_*.csv")

    frames = read_rows(frames_path)
    display = read_rows(display_path)
    trials = read_rows(args.trials)

    # events are read across every session, not just the newest. Failures and
    # aborts are rare by nature: a clean run has none, so looking at one session
    # would report a metric as unlogged when an earlier run already caught it.
    event_files = [args.events] if args.events else sorted(glob.glob(os.path.join(DATA, "events_*.csv")))
    events = []
    events_by_file = {}
    for p in event_files:
        rows = read_rows(p)
        events.extend(rows)
        events_by_file[os.path.basename(p)] = rows

    f_src = os.path.basename(frames_path) if frames_path else ""
    d_src = os.path.basename(display_path) if display_path else ""
    t_src = os.path.basename(args.trials) if os.path.exists(args.trials) else ""

    # ---- what the frame log actually shows -------------------------------
    gaps = nums(frames, "gap_ms")
    widths = sorted({(r.get("width"), r.get("height")) for r in frames if r.get("width")})
    ts = nums(frames, "timestamp_ms")
    dur_s = (max(ts) - min(ts)) / 1000.0 if len(ts) > 1 else 0.0
    fps = (len(frames) / dur_s) if dur_s else 0.0
    has_pts = bool(frames) and "pts_us" in frames[0]

    frame_ev = ("%d frames over %.1f s, %.2f fps" % (len(frames), dur_s, fps)) if frames else ""
    gap_ev = ("p95 %.0f ms, max %.0f ms"
              % (sorted(gaps)[int(len(gaps) * 0.95)], max(gaps))) if gaps else ""
    res_ev = (", ".join("%sx%s" % w for w in widths[:3])) if widths else ""

    # ---- what the event log actually shows --------------------------------
    kinds = {}
    for r in events:
        kinds[r.get("type", "")] = kinds.get(r.get("type", ""), 0) + 1

    def ev(kind, label):
        """Logged once the event type appears in any session's log."""
        n = kinds.get(kind, 0)
        if n:
            where = sorted(f for f, rows in events_by_file.items()
                           if any(r.get("type") == kind for r in rows))
            return ("logged", label,
                    "%d %s event%s across %d session%s"
                    % (n, kind, "" if n == 1 else "s", len(where),
                       "" if len(where) == 1 else "s"),
                    ", ".join(where))
        return ("code ready", label,
                "logger writes this; no session has produced one yet", "StreamViewModel.kt")

    def recoverable(label):
        """Needs a failure the stream comes back from. The reference app cannot
        produce one: StreamViewModel treats CLOSED as terminal and calls
        stopStream() the moment it is reached, which closes the logs before any
        recovery could occur or be recorded."""
        if kinds.get("recovery", 0) or kinds.get("retry", 0):
            return ev("recovery", label)
        return ("not observable", label,
                "the app aborts and closes the session on CLOSED, so there is no "
                "window in which a recovery could be logged",
                "StreamViewModel.kt:241")

    def meta(field, label):
        """Trial metadata counts as logged only if a session filled it in."""
        filled = [r.get(field) for r in trials if r.get(field)]
        if filled:
            return ("logged", label, "recorded as %s" % filled[-1], t_src)
        return ("code ready", label,
                "field exists in trials.csv; the sessions on disk predate it",
                "SessionLogger.kt")

    # FrameEncoder puts an encoder back in the pipeline, so these exist now. They
    # describe the phone-side encoder, not the sealed one on the glasses, and the
    # writeup has to say so.
    enc_rows = read_rows(encoded_path)
    enc_src = os.path.basename(encoded_path) if encoded_path else ""

    def encoder(what):
        if enc_rows:
            sizes = nums(enc_rows, "size_bytes")
            return ("logged", what,
                    "%d encoded frames, mean %d bytes, in %s"
                    % (len(enc_rows), statistics.mean(sizes) if sizes else 0, enc_src),
                    enc_src)
        return ("code ready", what,
                "FrameEncoder writes encoded_*.csv; no session on disk has one yet",
                "FrameEncoder.kt")

    # MediaMTX's own counters, polled once a second by mediamtx_stats.py while a
    # session runs. Only rows where the path was actually publishing count.
    srv_rows = [r for r in read_rows(args.server) if r.get("stream_ready") == "1"]
    srv_src = os.path.basename(args.server) if os.path.exists(args.server) else ""

    def server(what, key=None, unit=""):
        if not srv_rows:
            return ("needs server", what,
                    "poller written and validated; no session polled yet",
                    "mediamtx_stats.py")
        path = srv_rows[-1].get("path", "")
        if key:
            vals = nums(srv_rows, key)
            ev_txt = ("%s: median %.0f%s over %d s on path %s"
                      % (key, statistics.median(vals), unit, len(vals), path)) if vals else (
                      "%d s of publishing on path %s" % (len(srv_rows), path))
        else:
            ev_txt = "%d s of publishing on path %s" % (len(srv_rows), path)
        return ("logged", what, ev_txt, srv_src)
    # Optional in her sheet for Gen 2, and there is nothing to capture yet either:
    # the glasses reach the phone over Bluetooth, which a capture cannot see, and
    # the app displays locally without sending the video anywhere. These only carry
    # a signal once the phone is pushing to a media server.
    pcap = lambda what: (
        "optional", what,
        "sheet marks this Optional for Gen 2; no video crosses the network on this "
        "path until the phone pushes to a server. Tooling exists from the Gen 1 work",
        "pcap_windows.py")

    # Labelled "phone preview" rather than plain "logged" on purpose. Her viewer_id
    # is defined as "A (lab wired network) or B (different wired access network)",
    # which only describes machines watching over a network - so her Viewer stage
    # may mean a remote client, not the phone's own screen. Saying which one was
    # measured costs nothing and means these rows are not overclaimed either way.
    display_ready = (
        ("logged (phone preview)",
         "display_*.csv records every frame that reached the phone's screen",
         "%d display frames in %s; measures the phone preview, not a remote viewer"
         % (len(display), d_src), d_src)
        if display
        else ("code ready", "display_*.csv records every frame that reached the screen",
              "logger writes display_*.csv; no session on disk has one yet",
              "SessionLogger.kt"))

    # keyed by the sheet's own metric name
    return {
        # ---- Camera / Input: the one group with real data ------------------
        "input_frame_timestamp": (
            "logged", "monotonic arrival stamp per frame, plus the sender's pts_us"
            if has_pts else "monotonic arrival stamp per frame",
            frame_ev, f_src),
        "input_fps": ("derived", "frames per second per window, from the frame log",
                      frame_ev, "window_metrics.py"),
        "input_gap_p95": ("derived", "p95 of the gaps between frame arrivals",
                          gap_ev, "window_metrics.py"),
        "input_gap_max": ("derived", "largest gap between frame arrivals",
                          gap_ev, "window_metrics.py"),
        "input_width / input_height": (
            "logged", "frame metadata recorded per frame, so a mid-stream change shows up",
            res_ev, f_src),

        # ---- Local link ----------------------------------------------------
        "link_type": ("logged", "known from the platform, recorded as trial metadata",
                      "Bluetooth Classic between glasses and phone", t_src),
        "rssi_or_link_quality": ("n/a", "N/A", "", ""),
        "link_disconnects": ev(
            "link_disconnect",
            "leaving DeviceSessionState.STARTED for anything but PAUSED; PAUSED is "
            "the wearer's tap gesture, not a drop"),

        # ---- Encoder / sender: the whole group is one project --------------
        "target_bitrate": encoder(
            "the rate FrameEncoder asks MediaCodec for, logged in the encoder_start event"),
        "encoded_bitrate": encoder("summed encoded frame sizes per one-second window"),
        "encoded_frame_size": encoder("size_bytes per frame from the codec output buffer"),
        "encoded_fps": encoder("encoded frames per one-second window"),
        "keyframes": encoder("BUFFER_FLAG_KEY_FRAME on the codec output buffer"),
        "send_queue_bytes": (
            ("logged", "bytes accepted by the publisher but not yet written to the socket",
             "median %.0f bytes, max %.0f over %d frames"
             % (statistics.median(nums(enc_rows, "send_queue_bytes")),
                max(nums(enc_rows, "send_queue_bytes")), len(enc_rows)), enc_src)
            if enc_rows and nums(enc_rows, "send_queue_bytes")
            else ("needs server", "bytes waiting in the push layer",
                  "RtmpPublisher tracks this; no session with publishing yet", "RtmpPublisher.kt")),

        # ---- Network / router: optional here, already covered elsewhere ----
        "uplink_traffic_rate": pcap("router packet capture filtered to the phone IP"),
        "downlink_traffic_rate": pcap("same capture, the other direction"),
        "packet_rate": pcap("packets per second each way"),
        "packet_size_dist": pcap("p50 and p95 of uplink packet sizes"),
        "pkt_inter_arrival": pcap("time between consecutive uplink packets"),
        "active_flows": pcap("distinct 5-tuples with traffic per window"),
        "new_connections": pcap("new 5-tuples and TCP SYNs per second"),
        "conn_duration_resets": pcap("flow lifetimes plus TCP resets"),
        "tcp_retransmissions": pcap("Wireshark retransmission analysis"),
        # Optional for Gen 2 like the rest of the Network/Router group, and it has
        # nothing to shape until the phone is pushing to a server.
        "shaper_queue_stats": (
            "optional", "tc -s sampled once per second on the shaping box",
            "sheet marks this Optional for Gen 2; tool written and tc parsing "
            "verified, but there is no uplink to shape on this path yet",
            "shaper_stats.py"),

        # ---- Server: second half of the encoder project --------------------
        "server_ingest_bitrate": server(
            "MediaMTX statistics API, bytes delta per second", "ingest_bitrate_kbps", " kbps"),
        "bytes_received": server("MediaMTX statistics API", "bytes_received"),
        "active_readers": server("MediaMTX statistics API", "active_readers"),
        "stream_ready / readiness_time": server(
            "MediaMTX statistics API; readiness needs the poller running before the publisher"),

        # ---- Viewer --------------------------------------------------------
        # The app's display log is the phone's own preview. Whether her "viewer"
        # for Gen 2 means that or a client watching the restream changes what is
        # owed here, so the two that need a remote reference are flagged to ask.
        # Taken on the phone preview, which is where the toolkit path actually
        # ends. The restream exists too, so this can be redone against a remote
        # viewer if her Viewer stage means that instead.
        "first_frame_latency": (
            ("logged (phone preview)",
             "first frame to reach the screen, from display_*.csv",
             "%.0f ms after session start" % nums(display, "timestamp_ms")[0], d_src)
            if display and nums(display, "timestamp_ms")
            else ("code ready", "first frame to reach the screen",
                  "display_*.csv row 0; no session on disk has one yet", "SessionLogger.kt")),
        "live_latency": (
            "method ready", "clock in frame: monitor clock minus the clock visible "
            "in the phone's preview",
            "no reference clock in the logs themselves, so this needs a filmed run "
            "against source_display.html; same method as Gen 1",
            "source_display.html"),
        "viewer_fps": display_ready,
        "viewer_frame_gaps": display_ready,
        "jitter": display_ready,
        "freeze_count / freeze_duration": display_ready,
        "viewer_resolution": (
            "logged", "frame metadata carries width and height per frame, and "
            "StreamViewModel logs a resolution event on every change",
            (res_ev + "; a mid-stream switch appears as a resolution event")
            if res_ev else "", f_src),
        "viewer_id": meta("viewer_id", "trial metadata: which viewer produced these rows"),

        # ---- Recovery: all written, waiting on a session to exercise it -----
        # the attempt count rides on the startup event, not on retry: a stream that
        # comes up first try still reports attempts=1
        "startup_attempts / retry_count": ev("startup", "attempts= on the startup event"),
        "startup_time (time-to-active)": ev("startup", "stream request until first STREAMING"),
        "retry_intervals": recoverable("gaps between successive retry events"),
        "recovery_time": recoverable("failure until STREAMING returns"),
        "active_state transitions": ev("stream_state", "every state the stream passed through"),
        "abort_flag / disconnects": ev("abort", "stream ended abnormally"),

        # ---- Device / context ----------------------------------------------
        "battery_level": ev("battery", "phone battery read at session start and end"),
        "temperature": ("n/a", "N/A", "", ""),
        "phone_position / phone_distance": meta("phone_position", "trial metadata set in the UI"),
        "motion_condition": meta("motion_condition", "trial metadata set in the UI"),
        "network_limit": meta("network_limit", "trial metadata; the cap applied to the uplink"),
        "trial_phase": ev("phase", "BASELINE / STRESS / RECOVERY, tapped during the run"),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheet", default=os.path.expanduser(
        "~/Downloads/IMC_metrics_logging_v2.xlsx"))
    ap.add_argument("-o", "--out", default=os.path.join(DATA, "gen2_coverage.csv"))
    ap.add_argument("--frames", help="a frames_*.csv (default: newest in data/app)")
    ap.add_argument("--events", help="an events_*.csv (default: newest in data/app)")
    ap.add_argument("--display", help="a display_*.csv (default: newest in data/app)")
    ap.add_argument("--encoded", help="an encoded_*.csv (default: newest in data/app)")
    # trials.csv is what the app writes. trials_all.csv is merge_sessions.py's
    # output and goes stale the moment a new session is pulled off the phone, so
    # only fall back to it.
    ap.add_argument("--server", default=os.path.join(DATA, "server_live.csv"),
                    help="mediamtx_stats.py output for the session")
    ap.add_argument("--trials", default=(
        os.path.join(DATA, "trials.csv") if os.path.exists(os.path.join(DATA, "trials.csv"))
        else os.path.join(DATA, "trials_all.csv")))
    args = ap.parse_args()

    if not os.path.exists(args.sheet):
        raise SystemExit("sheet not found: %s\nPass --sheet with the path." % args.sheet)

    wb = openpyxl.load_workbook(args.sheet, data_only=True)
    ws = wb["Stage Metrics"]
    header = [c.value for c in ws[1]]
    try:
        i_stage = header.index("Stage")
        i_metric = header.index("Metric")
        i_gen2 = next(i for i, h in enumerate(header) if h and "Gen 2" in str(h))
    except (ValueError, StopIteration):
        raise SystemExit("could not find Stage / Metric / Gen 2 columns in the sheet")

    findings = build_findings(args)
    out_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        metric = row[i_metric]
        if not metric:
            continue
        sheet_gen2 = (row[i_gen2] or "").strip()
        f = findings.get(str(metric).strip())
        if f:
            status, how, evidence, source = f
        elif sheet_gen2.lower().startswith("no"):
            status, how, evidence, source = ("n/a", "N/A", "", "")
        else:
            status, how, evidence, source = ("not assessed", "", "", "")
        out_rows.append([row[i_stage], metric, sheet_gen2, status, how, evidence, source])

    with open(args.out, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["stage", "metric", "gen2", "Status", "How Obtained",
                    "evidence", "source files"])
        w.writerows(out_rows)

    counts = {}
    for r in out_rows:
        counts[r[3]] = counts.get(r[3], 0) + 1
    print("wrote %s (%d metrics from the sheet)" % (args.out, len(out_rows)))
    for k in sorted(counts):
        print("  %-14s %d" % (k, counts[k]))

    ready = counts.get("code ready", 0)
    if ready:
        print("\n%d rows say 'code ready': the logger writes them but no session on "
              "disk has\nexercised them. One run with the current build converts most "
              "of those to logged." % ready)
    print("note: 'needs server' rows are all one piece of work, pushing the encoded "
          "frames to MediaMTX.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

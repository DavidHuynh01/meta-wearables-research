"""Take the lab's metrics spreadsheet and add a column saying, for Meta Gen 1,
whether black-box testing can actually log each metric and what evidence exists.

Usage:
    python tools/blackbox/gen1_coverage.py --sheet "C:/path/IMC_metrics_logging_v2.xlsx"

Reads the Stage Metrics sheet so the metric names and order are exactly hers,
then joins on my findings. Numbers come out of the run files, so the evidence
column cannot drift from the data. Anything with no run yet says so rather than
being guessed.

my_status values:
    logged        real data in hand
    tool ready    tool built and validated, no run yet
    lab only      needs the router rig or a second viewer machine
    not loggable  cannot be obtained this way, with the reason
    n/a           the sheet already marks it No for Gen 1
"""

import argparse
import csv
import os
import statistics

DATA = os.path.join("data", "blackbox")

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl needed: python -m pip install openpyxl")


def read_rows(path):
    if not os.path.exists(path):
        return []
    with open(path, newline="") as f:
        return list(csv.DictReader(f))


def nums(rows, key):
    out = []
    for r in rows:
        try:
            out.append(float(r[key]))
        except (KeyError, TypeError, ValueError):
            pass
    return out


def med(vals, unit="", nd=2):
    if not vals:
        return ""
    return "median %.*f%s over %d windows" % (nd, statistics.median(vals), unit, len(vals))


def read_trials(trials_dir):
    """Every controller event log, as {trial_id: {event: ms}} plus phase counts."""
    out = []
    if not os.path.isdir(trials_dir):
        return out
    for name in sorted(os.listdir(trials_dir)):
        d = os.path.join(trials_dir, name)
        if not os.path.isdir(d):
            continue
        for fn in os.listdir(d):
            if not fn.endswith("_events.csv"):
                continue
            events, phases = {}, []
            for r in read_rows(os.path.join(d, fn)):
                try:
                    t = int(r["unix_time_ms"])
                except (KeyError, ValueError):
                    continue
                if r.get("event") == "phase":
                    phases.append((r.get("detail", ""), t))
                else:
                    events.setdefault(r.get("event", ""), t)
            out.append({"trial": name, "events": events, "phases": phases})
    return out


def build_findings(args):
    pcap = read_rows(args.pcap_windows)
    flows = read_rows(args.flows)
    viewer = read_rows(args.viewer_windows)
    lat = read_rows(args.latency)

    P = lambda k, u="", n=2: med(nums(pcap, k), u, n)
    V = lambda k, u="", n=1: med(nums(viewer, k), u, n)

    freezes = nums(viewer, "freeze_count")
    freeze_ms = nums(viewer, "freeze_ms")
    gap_max = nums(viewer, "gap_max_ms")
    first = nums(viewer, "first_frame_ms")
    lat_v = nums(lat, "latency_s")

    pcap_src = os.path.basename(args.pcap_windows)
    view_src = os.path.basename(args.viewer_windows)
    lat_src = os.path.basename(args.latency)

    # the detail-ratio series behind viewer_resolution. A tight spread is the
    # point: the instrument was calibrated against a synthetic 3x switch, which
    # moves the ratio ~30%, so noise of a few thousandths makes a null a real null
    # rather than an inconclusive one.
    sharp = nums(read_rows(args.sharpness), "detail_ratio")
    if sharp:
        spread = max(sharp) - min(sharp)
        sharpness_ev = ("no switch over %d samples; detail ratio %.3f-%.3f "
                        "(spread %.3f vs ~30%% for a real switch)"
                        % (len(sharp), min(sharp), max(sharp), spread))
    else:
        sharpness_ev = "tool written and validated against a synthetic switch; no run analysed yet"

    trials = read_trials(args.trials_dir)
    startups, aborts, phase_names = [], 0, set()
    for tr in trials:
        e = tr["events"]
        if "stream_request" in e and "viewer_A_first_frame" in e:
            startups.append((e["viewer_A_first_frame"] - e["stream_request"]) / 1000.0)
        if "abort" in e:
            aborts += 1
        phase_names.update(p for p, _ in tr["phases"])
    n_tr = len(trials)
    tr_src = ("%d trial event log%s" % (n_tr, "" if n_tr == 1 else "s")) if n_tr else ""

    def recovery(metric_default, evidence, needs):
        """Logged once a trial exists, otherwise still just a ready tool."""
        if n_tr:
            return ("logged", metric_default, evidence, tr_src)
        return ("tool ready", metric_default, needs, "trial_controller.py")

    # keyed by the sheet's own metric name
    return {
        "uplink_traffic_rate": ("logged", "router packet capture, bytes from the phone IP per second",
                                P("uplink_mbps", " Mbps"), pcap_src),
        "downlink_traffic_rate": ("logged", "same capture, bytes toward the phone IP",
                                  P("downlink_mbps", " Mbps"), pcap_src),
        "packet_rate": ("logged", "packets per second each way",
                        P("uplink_pkt_rate", " pkt/s", 0), pcap_src),
        "packet_size_dist": ("logged", "p50 and p95 of uplink packet sizes per second",
                             P("pkt_size_p50", " bytes", 0), pcap_src),
        "pkt_inter_arrival": ("logged", "time between consecutive uplink packets",
                              P("inter_arrival_p95_ms", " ms p95", 1), pcap_src),
        "active_flows": ("logged", "distinct 5-tuples with traffic per window",
                         ("%d flows across the session" % len(flows)) if flows else "", pcap_src),
        "new_connections": ("logged", "new 5-tuples and TCP SYNs per second, the repeated-setup signal",
                            P("new_connections", "", 1), pcap_src),
        "conn_duration_resets": ("logged", "flow lifetimes plus TCP resets",
                                 P("conn_resets", "", 1), pcap_src),
        "tcp_retransmissions": ("logged but empty",
                                "Wireshark retransmission analysis, TCP only",
                                "reads 0 because the Meta traffic is QUIC, whose recovery is not visible",
                                pcap_src),
        "shaper_queue_stats": ("lab only", "tc -s sampled once per second on the throttling router",
                               "tool written, tc parsing verified; only exists during a throttled trial",
                               "shaper_stats.py"),

        "first_frame_latency": ("logged", "first visible frame advance in the screen recording",
                                ("%.1f ms" % first[0]) if first else "", view_src),
        "live_latency": ("logged",
                         "clock in frame: viewer clock minus the source clock visible in the video",
                         (("median %.2f s over %d samples, spread %.2f s"
                           % (statistics.median(lat_v), len(lat_v), max(lat_v) - min(lat_v)))
                          if lat_v else ""), lat_src),
        "viewer_fps": ("logged", "screen recording analysed for frames where the picture advanced",
                       V("viewer_fps", " fps", 1), view_src),
        "viewer_frame_gaps": ("logged", "time between visible frame advances, p50/p95/max",
                              V("gap_p50_ms", " ms p50", 1), view_src),
        "jitter": ("logged", "spread of the viewer frame gaps",
                   V("gap_p95_ms", " ms p95", 1), view_src),
        "freeze_count / freeze_duration": ("logged",
                                           "runs with no visible advance over 500 ms, with 300 ms and 1 s as a sensitivity check",
                                           (("%d freezes, %.0f ms total, longest gap %.0f ms"
                                             % (sum(freezes), sum(freeze_ms), max(gap_max)))
                                            if freezes else ""), view_src),
        "viewer_resolution": ("logged (switch only)",
                              "per-frame detail ratio in the screen recording: upscaling suppresses the "
                              "finest gradients, so a resolution drop shows as a step even though the "
                              "player window never changes size",
                              sharpness_ev, "resolution_switch.py"),
        "viewer_id": ("lab only", "one recording per viewer, tagged A or B",
                      "needs a second viewer machine on a different access network", ""),

        "startup_attempts / retry_count": ("tool ready",
                                           "repeated startup trials; retries also show as extra connection setup in the capture",
                                           "needs a batch of startup trials", "trial_controller.py"),
        "startup_time (time-to-active)": recovery(
            "controller logs stream_request and the viewer's first frame",
            (("%.1f s from stream request to first frame at the viewer, %d trial%s"
              % (statistics.median(startups), len(startups),
                 "" if len(startups) == 1 else "s"))
             if startups else "trial ran but no first-frame mark"),
            "needs a trial with the controller running"),
        "retry_intervals": ("tool ready", "gaps between successive startup attempts",
                            "needs a batch of startup trials", "trial_controller.py"),
        "recovery_time": ("tool ready",
                          "stress_end until the metric is within 10% of its pre-stress median for 10 consecutive seconds",
                          "needs a throttled trial", "trial_controller.py"),
        "active_state transitions": recovery(
            "controller phase events plus flow changes in the capture",
            ("phases recorded: %s" % ", ".join(sorted(phase_names))) if phase_names else "",
            "needs a trial with the controller running"),
        "abort_flag / disconnects": recovery(
            "stream end events plus connection resets in the capture",
            ("%d abort%s across %d trial%s"
             % (aborts, "" if aborts == 1 else "s", n_tr, "" if n_tr == 1 else "s")),
            "needs a trial with the controller running"),

        "link_type": ("logged", "known from the platform, recorded as trial metadata",
                      "Bluetooth Classic between glasses and phone", ""),
        "battery_level": ("manual", "read off the devices and written into the trial metadata",
                          "phone side is easy; glasses battery has no API on this path", ""),
        "phone_position / phone_distance": ("logged", "trial metadata set by the operator",
                                            "carried in the trial id and the controller notes", "trial_controller.py"),
        "motion_condition": ("logged", "trial metadata set by the operator", "", "trial_controller.py"),
        "network_limit": ("logged", "trial metadata; the cap the controller applied",
                          "no cap in the runs so far, the ladder needs the lab router", "trial_controller.py"),
        "trial_phase": recovery(
            "controller writes BASELINE / STRESS / RECOVERY with timestamps",
            ("%d phase marks across %d trial%s"
             % (sum(len(t["phases"]) for t in trials), n_tr, "" if n_tr == 1 else "s")),
            "needs a trial with the controller running"),
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheet", default=os.path.expanduser(
        "~/Downloads/IMC_metrics_logging_v2.xlsx"))
    ap.add_argument("-o", "--out", default=os.path.join(DATA, "gen1_coverage.csv"))
    ap.add_argument("--pcap-windows", default=os.path.join(DATA, "phone_rehearsal_windows.csv"))
    ap.add_argument("--flows", default=os.path.join(DATA, "phone_rehearsal_flows.csv"))
    ap.add_argument("--viewer-windows", default=os.path.join(DATA, "run3_viewer_windows.csv"))
    ap.add_argument("--latency", default=os.path.join(DATA, "phone_rehearsal_latency.csv"))
    ap.add_argument("--sharpness", default=os.path.join(DATA, "run3_sharpness.csv"))
    ap.add_argument("--trials-dir", default=os.path.join("data", "trials"))
    args = ap.parse_args()

    if not os.path.exists(args.sheet):
        raise SystemExit("sheet not found: %s\nPass --sheet with the path." % args.sheet)

    wb = openpyxl.load_workbook(args.sheet, data_only=True)
    ws = wb["Stage Metrics"]
    header = [c.value for c in ws[1]]
    try:
        i_stage = header.index("Stage")
        i_metric = header.index("Metric")
        i_gen1 = next(i for i, h in enumerate(header) if h and "Gen 1" in str(h))
    except (ValueError, StopIteration):
        raise SystemExit("could not find Stage / Metric / Gen 1 columns in the sheet")

    findings = build_findings(args)
    n_trials = len(read_trials(args.trials_dir))
    trials_note = ("read %d controller event log%s from %s"
                   % (n_trials, "" if n_trials == 1 else "s", args.trials_dir)
                   ) if n_trials else ""
    out_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        metric = row[i_metric]
        if not metric:
            continue
        sheet_gen1 = (row[i_gen1] or "").strip()
        f = findings.get(str(metric).strip())
        if f:
            status, how, evidence, source = f
        elif sheet_gen1.lower().startswith("no"):
            # the sheet already rules these out: they happen inside the sealed
            # part of the pipeline, so there is nothing to obtain them with
            status, how, evidence, source = ("n/a", "N/A", "", "")
        else:
            status, how, evidence, source = ("not assessed", "", "", "")
        out_rows.append([row[i_stage], metric, sheet_gen1, status, how, evidence, source])

    with open(args.out, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["stage", "metric", "gen1", "Status", "How Obtained",
                    "evidence", "source files"])
        w.writerows(out_rows)

    counts = {}
    for r in out_rows:
        counts[r[3]] = counts.get(r[3], 0) + 1
    if trials_note:
        print(trials_note)
    print("wrote %s (%d metrics from the sheet)" % (args.out, len(out_rows)))
    for k in sorted(counts):
        print("  %-16s %d" % (k, counts[k]))
    print("\nnote: runs so far used a phone camera in place of the glasses, so "
          "this shows which columns are obtainable, not Gen 1 measurements.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
